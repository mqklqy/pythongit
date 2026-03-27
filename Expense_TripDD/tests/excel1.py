from openpyxl import load_workbook,Workbook

workbook = load_workbook(filename='test.xlsx')
print(workbook.sheetnames)

sheet = workbook['test']
print(sheet)

sheet = workbook.active
print(sheet)

print(sheet.dimensions)

cell = sheet['B1']
print(cell, cell.value)

col_content = sheet['B']
row_content = sheet[3]
print(row_content)

for row in row_content:
    print(row.value)

print('---------------------------------------------------')
for row in sheet.iter_rows(min_row=2, max_ro=5, min_col=1, max_col=4):
    for cell in row:
        print(cell.value, end=' ')
    print()


for i in range(ord('A'), ord('C')+1):
    print(i, chr(i))

#------------------------------------------------------------
workbook = load_workbook(filename='cosmetics.xlsx')
sheet = workbook.active

workbook_1 = Workbook()
sheet_1 = workbook_1.active

cells = sheet['D']
data_list = []
for cell in cells:
    if isinstance(cell.value, int) and cell.value > 100:
        data_list.append(cell.row)
print('输出满足条件的数据所在行数的列表：\n{}\n'.format(data_list))
j = 1
for row in data_list:
    for col in range(ord('A'), ord('G') + 1):
        sheet_1[chr(col) + str(j)] = sheet[chr(col) + str(row)].value
    print('正在写入第{}行数据'.format(j), end=' ')
    j += 1
workbook_1.save('cosmetics_other.xlsx')